import re
import pandas as pd
import openpyxl
import datetime
from pathlib import Path
from openpyxl.styles import PatternFill
from dateutil.relativedelta import *
from collections import Counter

if __name__ == "__main__":
    df123 = pd.read_excel(r"C:\Desktop\ПЦР 8603.xlsx")
    groups_by_tn123 = df123.groupby(["ТН"])
    tn_name123 = {}
    for key, group in groups_by_tn123:
        date_by_person = group["ФИО"]  ## Сортируем по ТН
        tn_name123[key] = date_by_person

    new_dict_tn123 = {}
    for gr in tn_name123:
        dates = []
        for date in tn_name123[gr]:
            dates.append(date)
        new_dict_tn123[gr] = list(set(dates))
    print(new_dict_tn123)

    new_dict_tn_123=[]
    for gr in new_dict_tn123:
        if len(new_dict_tn123[gr]) >2:
            for name in new_dict_tn123[gr]:
                new_dict_tn_123.append(name.rstrip())

    print(new_dict_tn_123)
    df = pd.read_excel(r"C:\Desktop\ПЦР 8603.xlsx")
    groups_by_fullname = df.groupby(["ФИО"]) ## Группируем по имени всех (собираем тесты одного человека в одноу группу)

    name_tn = {}
    for key, group in groups_by_fullname:
        date_by_person = group["ТН"] ## Сортируем по ТН
        name_tn[key.lower().rstrip()] = date_by_person

    new_dict_tn = {}
    for gr in name_tn:
        dates = []
        for date in name_tn[gr]:
            if date == 'нс':
                continue
            dates.append(date)
        new_dict_tn[gr.rstrip()] = dates

    print(f"new_dict_tn== {new_dict_tn}")
    dict = {}
    for key, group in groups_by_fullname:
        date_by_person = group["Дата фактическая"].sort_values() ## Сортируем по дате
        dict[key.lower().rstrip()] = date_by_person

    new_dict = {}
    for gr in dict:
        dates = []
        for date in dict[gr.rstrip()]:
            dates.append(date)
        new_dict[gr.rstrip()] = dates

    all_tests_dict = {}
    for name in new_dict:
        days = []
        for dates in range(len(new_dict[name.rstrip()])):
            d1 = new_dict[name.rstrip()][dates]
            d1 = d1.strftime("%d.%m.%Y")
            days.append(d1)
            new_name = re.sub(r'\s+', ' ', name.rstrip())
            all_tests_dict[new_name.rstrip()] = days

    print(f"all_tests_dict=== {all_tests_dict}") ## СПИСОК ВСЕХ ТЕСТОВ

    for bad_name in new_dict_tn_123:
        if bad_name.lower().rstrip() in all_tests_dict:
            del all_tests_dict[bad_name.lower().rstrip()]

    print(f"all_tests_dict=== {all_tests_dict}") ## СПИСОК ВСЕХ ТЕСТОВ

    #################### ПОЛУЧЕНИЕ ДИАПАЗОНОВ АНТИТЕЛ ###################
    df_anti = pd.read_excel(r"C:\Desktop\Антитела 8603.xlsx", skiprows=2)
    df_anti = df_anti[df_anti["Качественный результат"] == "положительный"]
    groups_by_fullname = df_anti.groupby(
        ["Фамилия", "Имя", "Отчество"])  ## Получаем людей только с положительным результатом

    dict = {}
    for key, group in groups_by_fullname:
        date_by_person = group["Дата результата"].sort_values()
        dict[key] = date_by_person

    new_dict = {}
    for gr in dict:
        dates = []
        for date in dict[gr]:
            dates.append(date)
        new_dict[gr] = dates

    res_dict_anti = {}
    for name in new_dict:  ## Составояем диапазон дат месяц результата теста + 3 месяца, получаем пары
        days = []
        for dates in range(len(new_dict[name])):
            d1 = new_dict[name][dates]
            date_plus_trhee_m = d1 + relativedelta(months=+3)
            d1 = d1.strftime("%d.%m.%Y")
            date_plus_trhee_m = date_plus_trhee_m.strftime("%d.%m.%Y")
            res_per = d1 + " - " + date_plus_trhee_m
            days.append(res_per)
        new_name = ""
        for combName in name:
            new_name += combName.lower() + " "

        res_dict_anti[new_name.rstrip()] = days

    ################### CHECK ANTI PER IN FAIL############################
    print("ANTITELA")
    dict_per_wht_anti = {}
    list_less_two = []

    for name in all_tests_dict.keys():
        for alive_name in res_dict_anti.keys():
            if name.rstrip().lower() == alive_name.rstrip().lower():  ## Если в списке нарушителей есть человек с антителами
                print(f"{alive_name}")
                test_dt = all_tests_dict[name]
                print(f"Per {test_dt}")

                anti_keys = list(res_dict_anti[alive_name.lower()])  ## Получаем диапазон антител
                anti_dt = []
                uniq_anti_keys = list(set(anti_keys))
                new_list = []
                if len(uniq_anti_keys) != len(anti_keys):
                    anti_keys = uniq_anti_keys

                for per in anti_keys:
                    dts = per.split(" - ")
                    for dt in dts:
                        anti_dt.append(dt)

                dates = [datetime.datetime.strptime(ts, "%d.%m.%Y") for ts in test_dt]
                dates.sort()
                sorted_dates = [datetime.datetime.strftime(ts, "%d.%m.%Y") for ts in dates]
                first_id = ''
                sec_id = ''

                correct_date = []
                num_of_pairs = len(anti_keys)
                for fail in test_dt:
                    for pairs in anti_keys:
                        count_of_pairs = 0
                        pars_date = pairs.split("-")
                        start_date = pars_date[0].strip()
                        end_date = pars_date[1].strip()
                        print(f"Day====={fail}")
                        print(f"start_date====={start_date}")
                        print(f"end_date====={end_date}")

                        if datetime.datetime.strptime(fail, "%d.%m.%Y") > datetime.datetime.strptime(start_date,
                                                                                                     "%d.%m.%Y") and datetime.datetime.strptime(
                                fail, "%d.%m.%Y") < datetime.datetime.strptime(end_date, "%d.%m.%Y"):
                            print("IM IN!")
                            break
                        else:
                            print("IM OUT!")
                            count_of_pairs = count_of_pairs + 1

                        if count_of_pairs == num_of_pairs:
                            correct_date.append(fail)

                print(correct_date)
                print()
                dict_per_wht_anti[name.rstrip()] = correct_date

    print(f"dict_per_wht_anti==== {dict_per_wht_anti}") ## ВСЕ ЛЮДИ С АНТИТЕЛАМИ

    dict_sorted_by_antitela = {}  ## ВСЕ ЛЮДИ БЕЗ АНТИТЕЛ
    for name_with_anti in all_tests_dict.keys():
        if name_with_anti not in dict_per_wht_anti.keys():
            dict_sorted_by_antitela[name_with_anti.rstrip()] = all_tests_dict[name_with_anti]

    dict_all_people_sorted_by_antitela = {**dict_sorted_by_antitela, **dict_per_wht_anti}  ## ВСЕ ЛЮДИ
    print(f"dict_all_people_sorted_by_antitela==== {dict_all_people_sorted_by_antitela}")

    ###################################### REMOTE WORK ##################################
    df_wt_remote = pd.read_excel(r"C:\Desktop\удаленка 2021.xlsx")
    groups_by_fullname = df_wt_remote.groupby(
        ["Табельный номер"])

    dict_of_remoted_work = {}
    for key, group in groups_by_fullname:
        list_of_start_dates = []
        list_of_end_dates = []
        periods = []
        start_date_remote_work = group["Начало"]
        end_date_remote_work = group["до какой даты оформлена заявка"]
        for start in start_date_remote_work:
            list_of_start_dates.append(start)
        for end in end_date_remote_work:
            list_of_end_dates.append(end)

        for per in range(len(list_of_start_dates)):
            start_date = list_of_start_dates[per].strftime("%d.%m.%Y")
            end_date = list_of_end_dates[per].strftime("%d.%m.%Y")
            periods.append(start_date + " - " + end_date)

        dict_of_remoted_work[key.lower().rstrip()] = periods

    new_dict_per_remote = {}
    for gr in dict_of_remoted_work:
        dates = []
        for date in dict_of_remoted_work[gr]:
            dates.append(date)
        new_dict_per_remote[gr] = dates

    print(f"PEOPLE ON REMOTE === {new_dict_per_remote}") ## ПЕРИОДЫ УДАЛЕНКИ

    ################# USE PERIOD REMOTE ##################################
    print()
    print("REMOTE")
    dict_of_remoted_work = {}
    for name in dict_all_people_sorted_by_antitela.keys():
        for remote_name in new_dict_per_remote.keys():
            if name.rstrip().lower() == remote_name.rstrip().lower():
                dates = dict_all_people_sorted_by_antitela[name]
                remote_keys = new_dict_per_remote[remote_name.rstrip().lower()]
                uniq_remote_per = list(set(remote_keys))
                new_list = []
                period_of_remote = []
                if len(uniq_remote_per) != len(list(set(remote_keys))):
                    remote_keys = uniq_remote_per

                date_non_per = []
                num_of_pairs = len(remote_keys)

                count_of_pairs = 0

                for fail in dates:
                    for pairs in remote_keys:
                        pars_date = pairs.split("-")
                        start_date = pars_date[0].strip()
                        end_date = pars_date[1].strip()

                        if datetime.datetime.strptime(fail, "%d.%m.%Y") > datetime.datetime.strptime(start_date,
                                                                                                     "%d.%m.%Y") and datetime.datetime.strptime(
                            fail, "%d.%m.%Y") < datetime.datetime.strptime(end_date, "%d.%m.%Y"):
                           
                            break
                        else:
                            count_of_pairs = count_of_pairs + 1


                        if count_of_pairs == num_of_pairs:
                            date_non_per.append(fail)
                    count_of_pairs = 0

                
                dict_of_remoted_work[name.rstrip()] = date_non_per ## ВСЕ ЛЮДИ, КОГО НЕ ОТМАЗАЛИ НИ АНТИТЕЛА НИ ДИСТАНТ


    print(f"dict_of_remoted_work==== {dict_of_remoted_work}")

    dict_sorted_by_remote= {}  ## ВСЕ ЛЮДИ БЕЗ ДИСТАНТА
    for name in dict_all_people_sorted_by_antitela.keys():
        if name not in dict_of_remoted_work.keys():
            dict_sorted_by_remote[name.rstrip()] = dict_all_people_sorted_by_antitela[name.rstrip()]

    print(f"dict_sorted_by_remote==== {dict_sorted_by_remote}")

    dict_all_people_sorted_by_antitela_and_remote = {**dict_sorted_by_remote, **dict_of_remoted_work}  ## ВСЕ ЛЮДИ
    print(f"dict_failer_non_anti_and_people_with_anti==== {dict_all_people_sorted_by_antitela_and_remote}")

    ################### GET WEEKEND ###################################################
    print()
    df_wt_weekend = pd.read_excel(r"C:\Desktop\отпуска с указанием страны.xlsx")
    groups_by_fullname = df_wt_weekend.groupby(
        ["Табельный номер"])

    dict_of_weekend_work = {}
    for key, group in groups_by_fullname:
        list_of_start_dates = []
        list_of_end_dates = []
        periods = []

        start_date_weekend = group["Начало"]
        end_date_weekend = group["Истечение"]
        for start in start_date_weekend:
            list_of_start_dates.append(start)
        for end in end_date_weekend:
            list_of_end_dates.append(end)

        for per in range(len(list_of_start_dates)):
            start_date = list_of_start_dates[per].strftime("%d.%m.%Y")
            end_date = list_of_end_dates[per].strftime("%d.%m.%Y")
            periods.append(start_date + " - " + end_date)

        dict_of_weekend_work[key.lower().rstrip()] = periods

    new_dict_per_weekend = {}
    for gr in dict_of_weekend_work:
        dates = []
        for date in dict_of_weekend_work[gr]:
            dates.append(date)
        new_dict_per_weekend[gr] = dates

    # ################# USE PERIOD WEEKEND ##################################
    print("WEEKEND")
    dict_of_weekend = {}
    for name in dict_all_people_sorted_by_antitela_and_remote.keys():
        for weekend_name in new_dict_per_weekend.keys():
            if name.rstrip().lower() == weekend_name.rstrip().lower():
                dates = dict_all_people_sorted_by_antitela_and_remote[name]
                weekwnd_keys = new_dict_per_weekend[weekend_name.rstrip().lower()]
                uniq_weekend_per = list(set(weekwnd_keys))
                new_list = []
                if len(uniq_weekend_per) != len(list(set(weekwnd_keys))):
                    weekwnd_keys = uniq_weekend_per

                date_non_per = []
                num_of_pairs = len(weekwnd_keys)
                count_of_pairs = 0
                for fail in dates:
                    for pairs in weekwnd_keys:
                        pars_date = pairs.split("-")
                        start_date = pars_date[0].strip()
                        end_date = pars_date[1].strip()

                        if datetime.datetime.strptime(fail, "%d.%m.%Y") > datetime.datetime.strptime(start_date,
                                                                                                     "%d.%m.%Y") and datetime.datetime.strptime(
                            fail, "%d.%m.%Y") < datetime.datetime.strptime(end_date, "%d.%m.%Y"):
                            
                            break
                        else:
                            count_of_pairs = count_of_pairs + 1


                        if count_of_pairs == num_of_pairs:
                            date_non_per.append(fail)
                    count_of_pairs = 0


                dict_of_weekend[name.rstrip()] = date_non_per


    dict_sorted_by_weekend = {}  ## ВСЕ ЛЮДИ БЕЗ ОТПУСКА
    for name in dict_all_people_sorted_by_antitela_and_remote.keys():
        if name not in dict_of_weekend.keys():
            dict_sorted_by_weekend[name.rstrip()] = dict_all_people_sorted_by_antitela_and_remote[name.rstrip()]


    dict_all_people_sorted_by_antitela_and_remote_and_weekend = {**dict_sorted_by_weekend, **dict_of_weekend}  ## ВСЕ ЛЮДИ

    ################### ПОДСЧЕТ РАЗНИЦЫ #######################################
    res_dict = {}

    for name in dict_all_people_sorted_by_antitela_and_remote_and_weekend:  ## Ищем пары дат между которыми разница больше 10
        days = {}

        for dates in range(len(dict_all_people_sorted_by_antitela_and_remote_and_weekend[name]) - 1):
            d1 = dict_all_people_sorted_by_antitela_and_remote_and_weekend[name][dates]
            d2 = dict_all_people_sorted_by_antitela_and_remote_and_weekend[name][dates + 1]
            d1 = datetime.datetime.strptime(d1, "%d.%m.%Y")
            d2 = datetime.datetime.strptime(d2, "%d.%m.%Y")
            if abs((d2 - d1).days) > 10:
                d1 = d1.strftime("%d.%m.%Y")
                d2 = d2.strftime("%d.%m.%Y")
                period = d1 + " - " + str(d2).replace("-", ".").split(" ")[0]
                d1 = datetime.datetime.strptime(d1, "%d.%m.%Y")
                d2 = datetime.datetime.strptime(d2, "%d.%m.%Y")
                fail_day = abs((d2 - d1).days)
                days[period] = fail_day
                res_dict[name.rstrip()] = days

    print()
    res_res_dict = res_dict
    res_dict_A = {}
    periods_of_failer_with_sub_periods_of_anti = {}
    for res_name in res_dict.keys():
        days = {}
        if res_name in res_dict_anti.keys():
            periods_fail = res_dict[res_name]
            periods_anti = res_dict_anti[res_name]
            for per_fail in periods_fail:
                for per_anti in periods_anti:
                    start_fail = datetime.datetime.strptime(per_fail.split(" - ")[0], "%d.%m.%Y")
                    end_fail = datetime.datetime.strptime(per_fail.split(" - ")[1], "%d.%m.%Y")
                    start_anti = datetime.datetime.strptime(per_anti.split(" - ")[0], "%d.%m.%Y")
                    end_anti = datetime.datetime.strptime(per_anti.split(" - ")[1], "%d.%m.%Y")
                    if start_anti >= start_fail and end_anti <= end_fail:
                        fail_day = abs((start_fail - end_fail).days)
                        anti_day = abs((start_anti - end_anti).days)
                        res = abs(fail_day-anti_day)
                        if res < 10:
                            days[per_fail] = res
                            res_dict_A[res_name.rstrip()] = days
                        else:
                            periods_of_failer_with_sub_periods_of_anti[res_name.rstrip()] = per_anti

    res_dic_B = {}
    periods_of_failer_with_sub_periods_of_remote = {}
    for res_name in res_dict.keys():
        days_B = {}
        if res_name in new_dict_per_remote.keys():
            periods_fail = res_dict[res_name]
            periods_remote = new_dict_per_remote[res_name]
            for per_fail in periods_fail:
                for per_remote in periods_remote:
                    start_fail = datetime.datetime.strptime(per_fail.split(" - ")[0], "%d.%m.%Y")
                    end_fail = datetime.datetime.strptime(per_fail.split(" - ")[1], "%d.%m.%Y")
                    start_remote = datetime.datetime.strptime(per_remote.split(" - ")[0], "%d.%m.%Y")
                    end_remote = datetime.datetime.strptime(per_remote.split(" - ")[1], "%d.%m.%Y")
                    if start_remote >= start_fail and end_remote <= end_fail:
                        fail_day = abs((start_fail - end_fail).days)
                        remote_day = abs((start_remote - end_remote).days)
                        res_remote = abs(fail_day - remote_day)
                        if res_remote < 10:
                            days_B[per_fail] = res_remote
                            res_dic_B[res_name.rstrip()] = days_B
                        else:
                            periods_of_failer_with_sub_periods_of_remote[res_name.rstrip()] = per_remote

    res_dic_C = {}
    periods_of_failer_with_sub_periods_of_weekend = {}
    for res_name in res_dict.keys():
        days_C={}
        if res_name in new_dict_per_weekend.keys():
            periods_fail = res_dict[res_name]
            periods_week = new_dict_per_weekend[res_name]
            for per_fail in periods_fail:
                for per_week in periods_week:
                    start_fail = datetime.datetime.strptime(per_fail.split(" - ")[0], "%d.%m.%Y")
                    end_fail = datetime.datetime.strptime(per_fail.split(" - ")[1], "%d.%m.%Y")
                    start_week = datetime.datetime.strptime(per_week.split(" - ")[0], "%d.%m.%Y")
                    end_week = datetime.datetime.strptime(per_week.split(" - ")[1], "%d.%m.%Y")
                    if start_week >= start_fail and end_week <= end_fail:
                        fail_day = abs((start_fail - end_fail).days)
                        week_day = abs((start_week - end_week).days)
                        res_week = abs(fail_day - week_day)
                        if res_week < 10:
                            days_C[per_fail] = res_week
                            res_dic_C[res_name.rstrip()] = days_C
                        else:
                            periods_of_failer_with_sub_periods_of_weekend[res_name.rstrip()] = per_week



    res_dict_with_subs_periods = {**periods_of_failer_with_sub_periods_of_remote, **periods_of_failer_with_sub_periods_of_weekend}
    res_ABC ={**res_dict_A, **res_dic_B, **res_dic_C}

    RESULT_RATATATATA = res_dict.copy()
    for fucking_name in res_dict.keys():
        for del_name in res_ABC.keys():
            if fucking_name == del_name:
                del RESULT_RATATATATA[fucking_name]

    dict_without_may = {}
    for ratata_name in RESULT_RATATATATA.keys():
        ratata_per = list(RESULT_RATATATATA[ratata_name].keys())
        new_per = []
        for per in ratata_per:
            start_fail = datetime.datetime.strptime(per.split(" - ")[0], "%d.%m.%Y")
            if not datetime.datetime.strptime(per.split(" - ")[0], "%d.%m.%Y") < datetime.datetime.strptime("13.05.2021", "%d.%m.%Y"):
                new_per.append(per)
        if len(new_per) != 0:
            dict_without_may[ratata_name.rstrip()] = new_per

    dict_non_blank_and_potential_intruder = {}
    for name_tn in new_dict_tn.keys():
        if len(new_dict_tn[name_tn]) == 0 and name_tn in dict_without_may.keys():
            del dict_without_may[name_tn]
            dict_non_blank_and_potential_intruder[name_tn] = {}

    new_dict_tn_without_blank = new_dict_tn.copy()
    for tn_name in new_dict_tn.keys():
        if len(new_dict_tn[tn_name]) == 0:
            del new_dict_tn_without_blank[tn_name]
        else:
            new_dict_tn_without_blank[tn_name] = new_dict_tn_without_blank[tn_name][0]

    ############################# GET ALL ENTER BY TN ###########################
    df_enter = pd.read_excel(r"C:\Desktop\vsp_useractivity_detail2.xlsx")
    groups_by_tn = df_enter.groupby(
        ["saphr_id"])  ## Группируем по имени всех (собираем тесты одного человека в одноу группу)

    dict_of_entered = {}
    for key, group in groups_by_tn:
        date_by_person = group["business_dt"]  ## Сортируем по ТН
        dict_of_entered[key] = list(date_by_person)

    dict_intruder_tn_and_periods = {}
    for name_tn in new_dict_tn_without_blank.keys():
        for intruder_name in dict_without_may.keys():
            if name_tn == intruder_name:
                dict_intruder_tn_and_periods[new_dict_tn_without_blank[name_tn]] = dict_without_may[intruder_name]

    dict_subs_tn_and_periods = {}
    for name_tn in new_dict_tn_without_blank.keys():
        for intruder_name in res_dict_with_subs_periods.keys():
            if name_tn == intruder_name:
                dict_subs_tn_and_periods[new_dict_tn_without_blank[name_tn]] = res_dict_with_subs_periods[intruder_name]

    ####################### ПРОВЕРЯЕМ ВХОЖДЕНИЯ В ПЕРИОДЫ ПЕРИОДОВ ВХОЖДЕНИЯ ПК ##############
    print()
    intruder_without_sub_period = {}
    for tn in dict_of_entered.keys():  ## ПЕРЕБОР ВХОЖДЕНИЙ
        for tn_per in dict_intruder_tn_and_periods.keys(): ## ПЕРЕБОР ПЕРИОДОВ НАРУШЕНИЙ
            if tn == tn_per:                                ## ЕСЛИ НАРУШИТЕЛЬ ВХОДИЛ В ПК
                list_of_enter_dates = dict_of_entered[tn]
                dates = []
                if tn in dict_subs_tn_and_periods.keys():     ## ЕСЛИ У НАРУШИТЕЛЯ ЕСТЬ ПЕРИОДЫ В ПЕРИОДАХ
                    print(f"--WITH SUB PERIODS--TN= {tn}")
                    sub_per_start = dict_subs_tn_and_periods[tn].split(" - ")[0] ## ПОЛУЧАЕМ НАЧАЛО ПОДПЕРИОДА
                    sub_per_end = dict_subs_tn_and_periods[tn].split(" - ")[1] ## ПОЛУЧАЕМ КОНЕЦ ПОДПЕРИОДА
                    per_fail = dict_intruder_tn_and_periods[tn_per] ## ПОЛУЧАЕМ ПЕРИОДЫ НАРУШЕНИЯ
                    print(f"tn== {tn} ====| sub_period==== {sub_per_start}-{sub_per_end}====|")
                    for per in per_fail:    ## ПЕРЕБЕРЕАМ ПЕРИОДЫ
                        start_fail = per.split(" - ")[0]  ## ПОЛУЧАЕМ НАЧАЛО ПЕРИОДА
                        end_fail = per.split(" - ")[1]    ## ПОЛУЧАЕМ КОНЕЦ ПЕРИОДА
                        print(f"Fail_period==== {start_fail}-{end_fail}====|")
                        for enter_date in list_of_enter_dates:
                            if (datetime.datetime.strptime(sub_per_start, "%d.%m.%Y") > datetime.datetime.strptime(start_fail, "%d.%m.%Y") and datetime.datetime.strptime(sub_per_end, "%d.%m.%Y") < datetime.datetime.strptime(end_fail, "%d.%m.%Y")) and (datetime.datetime.strptime(enter_date, "%d.%m.%Y") > datetime.datetime.strptime(sub_per_start, "%d.%m.%Y") and datetime.datetime.strptime(enter_date, "%d.%m.%Y") < datetime.datetime.strptime(sub_per_end, "%d.%m.%Y")):
                                print(f"Enter_date=== {enter_date}===| IN SUB PERIOD")
                                continue
                            elif datetime.datetime.strptime(sub_per_start, "%d.%m.%Y") > datetime.datetime.strptime(start_fail, "%d.%m.%Y") and datetime.datetime.strptime(sub_per_end, "%d.%m.%Y") < datetime.datetime.strptime(end_fail, "%d.%m.%Y") and not (datetime.datetime.strptime(enter_date, "%d.%m.%Y") > datetime.datetime.strptime(sub_per_start, "%d.%m.%Y") and datetime.datetime.strptime(enter_date, "%d.%m.%Y") < datetime.datetime.strptime(sub_per_end, "%d.%m.%Y")) and (datetime.datetime.strptime(enter_date, "%d.%m.%Y") > datetime.datetime.strptime(start_fail, "%d.%m.%Y") and datetime.datetime.strptime(enter_date, "%d.%m.%Y") < datetime.datetime.strptime(end_fail, "%d.%m.%Y")):
                                print(f"Enter_date=== {enter_date}===| HAS SUB PERIOD BUT DATE OUTSIDE IT")
                                dates.append(enter_date)
                            elif not (datetime.datetime.strptime(sub_per_start, "%d.%m.%Y") > datetime.datetime.strptime(start_fail, "%d.%m.%Y") and datetime.datetime.strptime(sub_per_end, "%d.%m.%Y") < datetime.datetime.strptime(end_fail, "%d.%m.%Y")) and (datetime.datetime.strptime(enter_date, "%d.%m.%Y") > datetime.datetime.strptime(start_fail, "%d.%m.%Y") and datetime.datetime.strptime(enter_date, "%d.%m.%Y") < datetime.datetime.strptime(end_fail, "%d.%m.%Y")):
                                print(f"Enter_date=== {enter_date}===| HASNT SUB PERIOD BUT DATE IN FAIL")
                                dates.append(enter_date)
                            else:
                                print(f"Enter_date=== {enter_date}===| NON PERIOD")

                else:
                    list_of_per = dict_intruder_tn_and_periods[tn_per]
                    print(f"--WITHOUT SUB PERIODS--TN= {tn}")
                    for per in list_of_per:
                        start_per = per.split(" - ")[0].rstrip()
                        end_per = per.split(" - ")[1].rstrip()
                        print(f"Fail_period==== {start_per}-{end_per}====|")
                        for date_enter in list_of_enter_dates:
                            if datetime.datetime.strptime(date_enter, "%d.%m.%Y") > datetime.datetime.strptime(start_per,
                                                                                                               "%d.%m.%Y") and datetime.datetime.strptime(
                                date_enter, "%d.%m.%Y") < datetime.datetime.strptime(end_per, "%d.%m.%Y"):
                                print(f"Enter_date=== {date_enter}===| IN PERIOD")
                                dates.append(date_enter)
                            else:
                                print(f"Enter_date=== {date_enter}===| NON IN PERIOD")

                print()
                intruder_without_sub_period[tn_per] = dates

    print(f"intruder_without_sub_period=== {intruder_without_sub_period}")


    ##PAINTING====================================================
    xlsx_writer = Path(r"C:\\Desktop\\vsp_useractivity_detail2.xlsx")
    wb = openpyxl.load_workbook(xlsx_writer)

    sheet = wb.active
    max_row = sheet.max_row
    max_col = sheet.max_column

    cells = sheet["A2":"G233367"]
    advanced_dir = {}
    for c1, c2, c3, c4, c5, c6, c7 in cells:
        for dic in intruder_without_sub_period.keys():
            dates = intruder_without_sub_period[dic]
            business_dt = c1.value
            cell_name = c3.value
            tn = c7.value
            for date in dates:
                if business_dt == date and str(tn).rstrip() == str(dic).rstrip():
                    print("FIND IT!!!")
                    sheet[c1.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c2.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c3.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c4.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c5.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c6.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
                    sheet[c7.coordinate].fill = PatternFill(fill_type="solid", start_color="FF8345")
    wb.save(r"C:\\Users\\19272773\\Desktop\\vsp_useractivity_detail2.xlsx")
    ################### CREATE EXCEL###################################################

    list_of_dict = [dict_without_may]
    period = []
    count_days = []
    namess = []
    for dict in list_of_dict:
        for name in dict.keys():
            for per_dic in dict[name]:
                period.append(per_dic)
                namess.append(name)

    df_with_anti_and_remote = pd.DataFrame()
    df_with_anti_and_remote["Name"] = namess
    df_with_anti_and_remote["Period"] = period
    df_with_anti = df_with_anti_and_remote.drop_duplicates(["Name", "Period"])

    list_of_df = []
    path = r"C:\\Desktop\\3.xlsx"
    writer = pd.ExcelWriter(path, engine='openpyxl')  # Результирующий файл

    list_of_df.append(df_with_anti_and_remote)
    list_of_name_sheet = ["SHeet1"]

    for df in range(len(list_of_df)):
        list_of_df[df].to_excel(writer, sheet_name=list_of_name_sheet[df])
        writer.save()
    writer.close()


